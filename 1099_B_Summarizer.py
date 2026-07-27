import io
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pypdf
import pdfplumber
import streamlit as st

class IRSTranscript1099BParser:
    def __init__(self):
        self.categories = [
            'Long-Term – Basis Reported',
            'Long-Term – Basis NOT Reported',
            'Short-Term – Basis Reported',
            'Short-Term – Basis NOT Reported',
            'Non-Covered / Unknown'
        ]

    def extract_text_from_pdf(self, pdf_file_bytes):
        full_text = ''
        try:
            reader = pypdf.PdfReader(pdf_file_bytes)
            for page in reader.pages:
                full_text += '\n' + (page.extract_text() or '')
        except Exception:
            with pdfplumber.open(pdf_file_bytes) as pdf:
                for page in pdf.pages:
                    full_text += '\n' + (page.extract_text() or '')
        return full_text

    def parse_transcript(self, raw_text):
        year_blocks = re.split(r'Tax Period Requested:\s*12-31-(20[12][0-9])', raw_text)
        all_records = []

        if len(year_blocks) > 1:
            for i in range(1, len(year_blocks), 2):
                year = year_blocks[i]
                block_text = year_blocks[i+1]
                self._extract_forms_from_block(year, block_text, all_records)
        else:
            self._extract_forms_from_block("All Years", raw_text, all_records)

        return pd.DataFrame(all_records)

    def _extract_forms_from_block(self, year, text, records):
        form_matches = list(re.finditer(r'Form\s*(1099-B|1099-DA)\b', text, flags=re.IGNORECASE))
        
        for idx, match in enumerate(form_matches):
            form_type = match.group(1).upper()
            start_pos = match.start()
            end_pos = form_matches[idx+1].start() if idx + 1 < len(form_matches) else len(text)
            form_text = text[start_pos:end_pos]

            company = self._extract_company_name(form_text, form_type)
            ein = self._extract_ein(form_text)

            # Extract fields
            d_match = re.search(
                r"Description:\s*(.*?)(?=\n[A-Z][A-Za-z\s]+:|\nSecond Notice|\nDate acquired|\nNoncovered|\nType of gain|$)",
                form_text,
                re.IGNORECASE | re.DOTALL
            )
            ds_match = re.search(r"Date\s*Sold\s*or\s*Disposed:\s*([0-9]{2}-[0-9]{2}-[0-9]{4}|[^\n]+)", form_text, re.IGNORECASE)
            da_match = re.search(r"Date\s*acquired:\s*([0-9]{2}-[0-9]{2}-[0-9]{4}|[^\n]+)", form_text, re.IGNORECASE)

            p_match = re.search(r"Proceeds:\s*\$?([0-9,]+(?:\.[0-9]{2})?)", form_text, re.IGNORECASE)
            b_match = re.search(r"(?:Cost or Basis|Cost|Basis):\s*\$?([0-9,]+(?:\.[0-9]{2})?)", form_text, re.IGNORECASE)
            w_match = re.search(r"Wash\s*Sale.*?\:\s*\$?([0-9,]+(?:\.[0-9]{2})?)", form_text, re.IGNORECASE)
            o_match = re.search(r"(?:Accrued Market Discount|Other Adjustment|1F)\s*:\s*\$?([0-9,]+(?:\.[0-9]{2})?)", form_text, re.IGNORECASE)

            description = " ".join(d_match.group(1).split()) if d_match else "N/A"
            date_sold = ds_match.group(1).strip() if ds_match else "N/A"
            date_acquired = da_match.group(1).strip() if da_match else "N/A"

            proceeds = float(p_match.group(1).replace(',', '')) if p_match else 0.0
            has_explicit_basis = True if b_match else False
            basis = float(b_match.group(1).replace(',', '')) if b_match else 0.0
            wash = float(w_match.group(1).replace(',', '')) if w_match else 0.0
            other = float(o_match.group(1).replace(',', '')) if o_match else 0.0

            is_empty_form = (proceeds == 0.0 and basis == 0.0)

            category = self._classify_category(form_text, has_explicit_basis)

            # Format basis: if 0.0, keep as blank string ""
            formatted_basis = "" if basis == 0.0 else basis

            records.append({
                'Year': str(year),
                'Company': company,
                'EIN': ein,
                'Form Type': form_type,
                'Description': description,
                'Date Sold or Disposed': date_sold,
                'Date Acquired': date_acquired,
                'Category': category,
                'Proceeds': proceeds,
                'Cost Basis': formatted_basis,
                'Wash Sale': wash,
                'Other Adjustments': other,
                'Has Explicit Basis': has_explicit_basis,
                'Is Empty Form': is_empty_form
            })

    def _extract_company_name(self, form_text, form_type):
        lines = [l.strip() for l in form_text.split('\n') if l.strip()]
        for i, line in enumerate(lines):
            if re.search(r"(Payer's Federal Identification Number|Filer's TIN|Payer:|Filer:)", line, re.IGNORECASE):
                for offset in range(1, 4):
                    if i + offset < len(lines):
                        cand = lines[i + offset].strip()
                        cand_clean = re.sub(r'^(Payer|Filer|FIN|TIN|EIN|Recipient|Taxpayer|Page|P\b).*', '', cand, flags=re.IGNORECASE).strip()
                        if len(cand_clean) > 2 and not cand_clean.isdigit():
                            broker_name = cand_clean.upper()
                            return f"{broker_name} (1099-DA)" if form_type == "1099-DA" else broker_name
        
        return "UNKNOWN DIGITAL ASSET BROKER (1099-DA)" if form_type == "1099-DA" else "UNKNOWN BROKERAGE"

    def _extract_ein(self, form_text):
        match = re.search(r"(?:Payer's Federal Identification Number|Filer's TIN|FIN|TIN|EIN)\s*(?:\([A-Z]+\))?\s*:\s*([0-9]{2}-[0-9]{7}|[0-9]{9})", form_text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
        
        fallback = re.search(r"\b([0-9]{2}-[0-9]{7})\b", form_text)
        return fallback.group(1).strip() if fallback else "N/A"

    def _classify_category(self, form_text, has_explicit_basis):
        f = ' '.join(form_text.upper().split())
        
        if ("CANNOT DETERMINE WHETHER THE RECIPIENT SHOULD CHECK BOX B OR BOX E" in f or 
            "HOLDING PERIOD IS UNKNOWN" in f or 
            "NONCOVERED SECURITY BASIS NOT REPORTED TO IRS" in f or
            "HOLDING PERIOD UNKNOWN" in f):
            
            if "SHORT TERM TRANSACTION" not in f and "LONG TERM TRANSACTION" not in f:
                return "Non-Covered / Unknown"

        is_long = "LONG-TERM" in f or "LONG TERM" in f or "BOX B" in f or "BOX E" in f
        is_short = "SHORT-TERM" in f or "SHORT TERM" in f or "BOX A" in f or "BOX D" in f

        basis_reported = has_explicit_basis
        if "BASIS IS NOT BEING REPORTED" in f or "BASIS NOT REPORTED" in f or "NONCOVERED SECURITY BASIS NOT REPORTED" in f:
            basis_reported = False
        elif "BASIS IS BEING REPORTED" in f or "BASIS REPORTED TO IRS" in f:
            basis_reported = True

        if is_long:
            return "Long-Term – Basis Reported" if basis_reported else "Long-Term – Basis NOT Reported"
        elif is_short:
            return "Short-Term – Basis Reported" if basis_reported else "Short-Term – Basis NOT Reported"
        else:
            return "Non-Covered / Unknown"

    def generate_proseries_summary(self, df):
        if df.empty:
            return pd.DataFrame()

        summary_rows = []
        years = sorted(df['Year'].unique(), reverse=True)
        
        for yr in years:
            yr_df = df[df['Year'] == yr]
            for comp in sorted(yr_df['Company'].unique()):
                comp_df = yr_df[yr_df['Company'] == comp]
                for cat in self.categories:
                    cat_df = comp_df[comp_df['Category'] == cat]
                    forms_count = len(cat_df)
                    
                    # Convert cost basis to float for summation safely
                    basis_sum = sum(float(x) for x in cat_df['Cost Basis'] if str(x).replace('.','',1).isdigit() and float(x) > 0)
                    proceeds_sum = cat_df['Proceeds'].sum()
                    empty_forms_count = int(cat_df['Is Empty Form'].sum()) if forms_count > 0 else 0

                    formatted_basis = "" if basis_sum == 0.0 else basis_sum

                    if forms_count > 0 and not (proceeds_sum == 0 and basis_sum == 0):
                        summary_rows.append({
                            'Tax Year': yr,
                            'Company': comp,
                            'Category': cat,
                            'Proceeds': proceeds_sum,
                            'Cost Basis': formatted_basis,
                            'Wash Sale Disallowed': cat_df['Wash Sale'].sum(),
                            'Other Adjustments/Income': cat_df['Other Adjustments'].sum(),
                            'Total Forms': forms_count,
                            'Empty Forms': empty_forms_count
                        })
                    else:
                        summary_rows.append({
                            'Tax Year': yr,
                            'Company': comp,
                            'Category': cat,
                            'Proceeds': 'None',
                            'Cost Basis': '',
                            'Wash Sale Disallowed': 'None',
                            'Other Adjustments/Income': 'None',
                            'Total Forms': forms_count,
                            'Empty Forms': empty_forms_count
                        })
        return pd.DataFrame(summary_rows)

    def generate_proseries_ready(self, df):
        if df.empty:
            return pd.DataFrame()

        ready_rows = []
        years = sorted(df['Year'].unique(), reverse=True)

        category_group_names = {
            'Short-Term – Basis Reported': 'SHORT TERM GROUP',
            'Short-Term – Basis NOT Reported': 'SHORT TERM GROUP',
            'Long-Term – Basis Reported': 'LONG TERM GROUP',
            'Long-Term – Basis NOT Reported': 'LONG TERM GROUP',
            'Non-Covered / Unknown': 'NONCOVERED GROUP'
        }

        for yr in years:
            yr_df = df[df['Year'] == yr]
            for comp in sorted(yr_df['Company'].unique()):
                comp_df = yr_df[yr_df['Company'] == comp]
                ein_val = comp_df['EIN'].iloc[0] if 'EIN' in comp_df.columns and not comp_df['EIN'].empty else "N/A"
                
                for cat in self.categories:
                    cat_df = comp_df[comp_df['Category'] == cat]
                    forms_count = len(cat_df)

                    basis_sum = sum(float(x) for x in cat_df['Cost Basis'] if str(x).replace('.','',1).isdigit() and float(x) > 0)
                    proceeds_sum = cat_df['Proceeds'].sum()

                    if forms_count > 0 and not (proceeds_sum == 0 and basis_sum == 0):
                        if forms_count > 1:
                            # Group entry
                            desc_name = category_group_names.get(cat, 'GROUP ENTRY')
                            dt_sold = f"12/31/{yr}" if yr.isdigit() else "12/31/2025"
                            dt_acquired = "VARIOUS"
                        else:
                            # Single form entry
                            form_desc = str(cat_df['Description'].iloc[0]).strip()
                            upper_desc = form_desc.upper()

                            # Non-covered specific rule check
                            if cat == 'Non-Covered / Unknown':
                                if (not form_desc or 
                                    upper_desc in ['N/A', 'NONE', 'SEE DETAIL STATEMENT'] or 
                                    upper_desc.startswith('NONCOVERED')):
                                    desc_name = "NONCOVERED SECURITY"
                                else:
                                    desc_name = form_desc
                            else:
                                if upper_desc in ['SEE DETAIL STATEMENT'] or upper_desc.startswith('NONCOVERED'):
                                    desc_name = "NONCOVERED SECURITY"
                                else:
                                    desc_name = form_desc if form_desc and upper_desc not in ['N/A', 'NONE'] else category_group_names.get(cat, cat)
                            
                            dt_sold = str(cat_df['Date Sold or Disposed'].iloc[0]).strip()
                            raw_acq = str(cat_df['Date Acquired'].iloc[0]).strip()
                            
                            if not raw_acq or raw_acq in ['00-00-0000', 'N/A', 'NONE', '00/00/0000']:
                                dt_acquired = "VARIOUS"
                            else:
                                dt_acquired = raw_acq

                        formatted_basis = "" if basis_sum == 0.0 else basis_sum

                        ready_rows.append({
                            'Tax Year': yr,
                            'Company': comp,
                            'EIN': ein_val,
                            'Description / Group Name': desc_name,
                            'Category Class': cat,
                            'Date Sold': dt_sold,          # Date Sold BEFORE Date Acquired
                            'Date Acquired': dt_acquired,
                            'Proceeds': proceeds_sum,
                            'Cost Basis': formatted_basis, # Blank if 0
                            'Wash Sale Disallowed': cat_df['Wash Sale'].sum(),
                            'Other Adjustments': cat_df['Other Adjustments'].sum(),
                            'Form Count': forms_count
                        })
                        
        return pd.DataFrame(ready_rows)

def export_to_proseries_excel(summary_df, raw_df, ready_df, output_stream):
    wb = openpyxl.Workbook()
    
    font_title = Font(name='Calibri', size=16, bold=True, color='1F4E78')
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    font_year_header = Font(name='Calibri', size=12, bold=True, color='1F4E78')
    fill_year_header = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    # -------------------------------------------------------------
    # SHEET 1: ProSeries Ready (Date Sold BEFORE Date Acquired)
    # -------------------------------------------------------------
    ws_ready = wb.active
    ws_ready.title = "ProSeries Ready"
    ws_ready.views.sheetView[0].showGridLines = True

    ws_ready.append(["IRS Wage & Income Transcript - ProSeries Ready Import/Entry Sheet"])
    ws_ready.cell(row=1, column=1).font = font_title
    ws_ready.append(["This tool is developed by Md. Mamun Sarder"])
    ws_ready.append([])

    ready_headers = ["Tax Year", "Company / Brokerage", "EIN", "Description / Group Name", "Category", "Date Sold", "Date Acquired", "Proceeds", "Cost Basis", "Wash Sale Disallowed", "Other Adjustments", "Form Count"]
    ws_ready.append(ready_headers)
    for col in range(1, len(ready_headers) + 1):
        cell = ws_ready.cell(row=4, column=col)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row_idx = 5
    current_year = None
    if not ready_df.empty:
        for _, row in ready_df.iterrows():
            if row['Tax Year'] != current_year:
                current_year = row['Tax Year']
                ws_ready.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=12)
                banner_cell = ws_ready.cell(row=row_idx, column=1, value=f"TAX YEAR: {current_year}")
                banner_cell.font = font_year_header
                banner_cell.fill = fill_year_header
                row_idx += 1

            ws_ready.append([
                row['Tax Year'],
                row['Company'],
                row['EIN'],
                row['Description / Group Name'],
                row['Category Class'],
                row['Date Sold'],
                row['Date Acquired'],
                row['Proceeds'],
                row['Cost Basis'], # Displays as blank if empty
                row['Wash Sale Disallowed'],
                row['Other Adjustments'],
                row['Form Count']
            ])
            for col in range(1, 13):
                cell = ws_ready.cell(row=row_idx, column=col)
                cell.border = thin_border
                if col in [8, 9, 10, 11]:
                    if isinstance(cell.value, (int, float)) and cell.value != "":
                        cell.number_format = '$#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.alignment = Alignment(horizontal='center')
                elif col in [1, 3, 6, 7, 12]:
                    cell.alignment = Alignment(horizontal='center')
            row_idx += 1

    for col in ws_ready.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_ready.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # -------------------------------------------------------------
    # SHEET 2: Summary
    # -------------------------------------------------------------
    ws1 = wb.create_sheet(title="Summary")
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.append(["IRS Wage & Income Transcript - 1099-B & 1099-DA Summary"])
    ws1.cell(row=1, column=1).font = font_title
    ws1.append(["This tool is developed by Md. Mamun Sarder"])
    ws1.append([])
    
    headers = ["Tax Year", "Company / Brokerage", "Category", "Proceeds", "Cost Basis", "Wash Sale Disallowed", "Other Adjustments/Income", "Form Count", "Empty Forms"]
    ws1.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=4, column=col)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    row_idx = 5
    current_year = None
    for _, row in summary_df.iterrows():
        if row['Tax Year'] != current_year:
            current_year = row['Tax Year']
            ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
            banner_cell = ws1.cell(row=row_idx, column=1, value=f"TAX YEAR: {current_year}")
            banner_cell.font = font_year_header
            banner_cell.fill = fill_year_header
            row_idx += 1
            
        ws1.append([
            row['Tax Year'],
            row['Company'],
            row['Category'],
            row['Proceeds'],
            row['Cost Basis'],
            row['Wash Sale Disallowed'],
            row['Other Adjustments/Income'],
            row['Total Forms'],
            row['Empty Forms']
        ])
        for col in range(1, 10):
            cell = ws1.cell(row=row_idx, column=col)
            cell.border = thin_border
            if col in [4, 5, 6, 7]:
                if isinstance(cell.value, (int, float)) and cell.value != "":
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='center')
            elif col in [8, 9]:
                cell.alignment = Alignment(horizontal='center')
        row_idx += 1
        
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # -------------------------------------------------------------
    # SHEET 3+: Detail [YEAR]
    # -------------------------------------------------------------
    if not raw_df.empty:
        years_in_data = sorted(raw_df['Year'].unique(), reverse=True)
        raw_headers = list(raw_df.columns)
        
        for yr in years_in_data:
            yr_raw_df = raw_df[raw_df['Year'] == yr]
            if yr_raw_df.empty:
                continue
                
            sheet_title = f"Detail {yr}"
            ws_yr = wb.create_sheet(title=sheet_title)
            ws_yr.views.sheetView[0].showGridLines = True
            
            ws_yr.append(raw_headers)
            for col in range(1, len(raw_headers) + 1):
                cell = ws_yr.cell(row=1, column=col)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for _, r in yr_raw_df.iterrows():
                ws_yr.append(list(r))
                
            for col in ws_yr.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws_yr.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(output_stream)

# --- Streamlit Web Application Interface ---
st.set_page_config(page_title="1099-B Summarizer", layout="wide")
st.title("📄 IRS Wage & Income Transcript 1099-B & 1099-DA Summarizer")
st.caption("🛠️ **This tool is developed by Md. Mamun Sarder**")
st.write("Upload an IRS PDF transcript to generate a ProSeries-ready Excel reconciliation workbook.")

uploaded_file = st.file_uploader("Upload PDF Transcript", type=["pdf"])

if uploaded_file is not None:
    if st.button("🚀 Process & Generate Excel Summary"):
        with st.spinner("Analyzing transcript and extracting 1099-B / 1099-DA records..."):
            parser = IRSTranscript1099BParser()
            pdf_bytes = io.BytesIO(uploaded_file.getvalue())
            raw_text = parser.extract_text_from_pdf(pdf_bytes)
            
            raw_df = parser.parse_transcript(raw_text)
            summary_df = parser.generate_proseries_summary(raw_df)
            ready_df = parser.generate_proseries_ready(raw_df)
            
            excel_buffer = io.BytesIO()
            export_to_proseries_excel(summary_df, raw_df, ready_df, excel_buffer)
            excel_buffer.seek(0)
            
            st.success("Analysis Complete!")
            
            # --- WEB OUTPUT: EXACT PROSERIES READY MATRIX ---
            st.subheader("📊 ProSeries Ready Input Matrix")
            
            if not ready_df.empty:
                display_ready_df = ready_df.copy()
                for col in ['Proceeds', 'Cost Basis', 'Wash Sale Disallowed', 'Other Adjustments']:
                    display_ready_df[col] = display_ready_df[col].apply(
                        lambda x: f"${x:,.2f}" if isinstance(x, (int, float)) and x != "" else (str(x) if x != "" else "")
                    )
                
                st.dataframe(display_ready_df, use_container_width=True)
            else:
                st.info("No active transactions found with positive proceeds or cost basis.")

            # --- FULL EXCEL DOWNLOAD ---
            st.download_button(
                label="📥 Download Full ProSeries Excel Summary",
                data=excel_buffer,
                file_name=f"ProSeries_{uploaded_file.name.replace('.pdf', '')}_Summary.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

st.markdown("---")
st.markdown("<p style='text-align: center; color: gray;'>Designed for ProSeries Tax Software | Developed by <b>Md. Mamun Sarder</b></p>", unsafe_allow_html=True)
